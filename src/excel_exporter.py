"""Write the ``Debit Memo Results`` sheet into the uploaded short-paid workbook.

The original workbook is loaded with openpyxl and a NEW sheet is appended; no
existing worksheet, value, formula, or formatting is touched. (pandas.ExcelWriter
is deliberately avoided because it rewrites the whole workbook and would destroy
formatting and other features.)
"""
from __future__ import annotations

import io

import openpyxl

from .formatting import RESULT_COLUMNS, style_result_sheet
from .models import DebitMemoResult

BASE_SHEET_NAME = "Debit Memo Results"


def unique_sheet_name(existing: list[str], base: str = BASE_SHEET_NAME) -> str:
    """Return ``base`` or the next free ``base 2``, ``base 3`` ... name."""
    if base not in existing:
        return base
    n = 2
    while f"{base} {n}" in existing:
        n += 1
    return f"{base} {n}"


def _row_values(result: DebitMemoResult) -> list:
    return [
        result.vendor_reference or "",
        result.debit_number or "",
        result.po_number or "",
        result.cell_text or "",
        result.status,
        result.review_reason or "",
    ]


def build_workbook(sp_bytes: bytes, results: list[DebitMemoResult]) -> bytes:
    """Return new workbook bytes: every original sheet plus the results sheet."""
    wb = openpyxl.load_workbook(io.BytesIO(sp_bytes))  # keeps formulas & formatting
    sheet_name = unique_sheet_name(list(wb.sheetnames))
    ws = wb.create_sheet(title=sheet_name)  # appended at the end

    ws.append(RESULT_COLUMNS)
    for result in results:
        ws.append(_row_values(result))

    style_result_sheet(ws, n_data_rows=len(results))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
