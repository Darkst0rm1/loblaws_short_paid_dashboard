"""Workbook loading helpers (LCL database and Short Paid List)."""

from __future__ import annotations

import hashlib
import io
import logging

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def list_sheets(file_bytes: bytes) -> list[str]:
    """Return the worksheet names of an .xlsx/.xlsm workbook."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=False, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_sheet(file_bytes: bytes, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    """Read a worksheet into a DataFrame as strings-preserving as practical.

    ``dtype=object`` keeps UPCs from being coerced to float (which would add
    a trailing ``.0`` / scientific notation).
    """
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row,
        dtype=object,
        engine="openpyxl",
    )
    df.columns = [str(c).strip() for c in df.columns]
    return df


def choose_lcl_sheet(sheet_names: list[str], preferred: str = "Material listing 04.29.26") -> str | None:
    """Pick the LCL worksheet: preferred name, else first sheet starting with
    'Material listing', else None (caller searches by columns)."""
    if preferred in sheet_names:
        return preferred
    for name in sheet_names:
        if name.lower().startswith("material listing"):
            return name
    return None


def file_hash(file_bytes: bytes) -> str:
    """SHA-256 hash for duplicate-file detection."""
    return hashlib.sha256(file_bytes).hexdigest()
