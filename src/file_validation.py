"""Upload validation and worksheet selection for the two Excel inputs.

* Material list: ``.xlsx`` / ``.xlsm``; default sheet ``Material listing
  04.29.26``, otherwise the first sheet whose header row has a ``Material``
  column.
* Short-paid list: ``.xlsx``; default sheet ``06.08.26``, otherwise the first
  sheet whose header row contains one of the reference aliases.
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from .reference_matcher import REFERENCE_COLUMN_ALIASES

MATERIAL_EXTENSIONS = (".xlsx", ".xlsm")
SHORT_PAID_EXTENSIONS = (".xlsx",)

DEFAULT_MATERIAL_SHEET = "Material listing 04.29.26"
DEFAULT_SHORT_PAID_SHEET = "06.08.26"
MATERIAL_COLUMN = "Material"


def has_valid_extension(filename: str, extensions) -> bool:
    return bool(filename) and filename.lower().endswith(tuple(extensions))


def list_sheets(file_bytes: bytes) -> list[str]:
    """Return worksheet names without loading cell data."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def read_sheet(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Read one worksheet as text-preserving objects (header in row 1)."""
    df = pd.read_excel(
        io.BytesIO(file_bytes), sheet_name=sheet_name, dtype=object, engine="openpyxl"
    )
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _header_of(file_bytes: bytes, sheet_name: str) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    return [str(c).strip() for c in header_row if c is not None]


def choose_material_sheet(file_bytes: bytes, sheets: list[str]) -> str:
    """Default to the known material sheet, else the first with a Material column."""
    if DEFAULT_MATERIAL_SHEET in sheets:
        return DEFAULT_MATERIAL_SHEET
    for name in sheets:
        try:
            header = [h.lower() for h in _header_of(file_bytes, name)]
        except Exception:
            continue
        if MATERIAL_COLUMN.lower() in header:
            return name
    return sheets[0] if sheets else ""


def choose_short_paid_sheet(file_bytes: bytes, sheets: list[str]) -> str:
    """Default to the known short-paid sheet, else the first with a ref alias."""
    if DEFAULT_SHORT_PAID_SHEET in sheets:
        return DEFAULT_SHORT_PAID_SHEET
    aliases = {a.lower() for a in REFERENCE_COLUMN_ALIASES}
    for name in sheets:
        try:
            header = {h.lower() for h in _header_of(file_bytes, name)}
        except Exception:
            continue
        if header & aliases:
            return name
    return sheets[0] if sheets else ""
