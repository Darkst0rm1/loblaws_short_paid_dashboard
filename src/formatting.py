"""Text and worksheet formatting helpers for the debit-memo result sheet.

Keeps all presentation rules in one place:

* ``format_money`` / ``format_quantity`` / ``format_item_line`` build the text
  that goes inside the combined "Material, Quantity # and Amount" cell.
* ``style_result_sheet`` applies the openpyxl formatting (bold header, freeze,
  autofilter, wrap, top-align, column widths, Manual-Review highlight).
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import STATUS_REVIEW

# Currency format used inside generated text (and as the cell number format hint).
CURRENCY_FORMAT = "$#,##0.00"

# Result-sheet column order (must match excel_exporter / processor output).
RESULT_COLUMNS = [
    "Vendor Reference Number",
    "Debit Number",
    "PO Number",
    "Material, Quantity # and Amount",
    "Status",
    "Review Reason",
]

# Practical column widths (the Material/Qty/Amount column is intentionally wide).
_COLUMN_WIDTHS = {
    "Vendor Reference Number": 22,
    "Debit Number": 16,
    "PO Number": 16,
    "Material, Quantity # and Amount": 46,
    "Status": 16,
    "Review Reason": 52,
}

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True)
_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # soft amber for Manual Review
_TOP_WRAP = Alignment(vertical="top", wrap_text=True)
_TOP_WRAP_HEADER = Alignment(vertical="center", wrap_text=True)


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def format_money(value) -> str:
    """Render a money value as ``$#,##0.00`` (e.g. ``$6,375.60``)."""
    d = _to_decimal(value)
    if d is None:
        return "$0.00"
    return "${:,.2f}".format(d.quantize(Decimal("0.01")))


def format_quantity(value) -> str:
    """Render a quantity, dropping a trailing ``.0`` (70 not 70.0)."""
    d = _to_decimal(value)
    if d is None:
        return ""
    if d == d.to_integral_value():
        return str(int(d))
    return str(d.normalize())


def format_item_line(label: str, qty, total) -> str:
    """Build one item line: ``{label}, {qty}, ${total}``.

    ``label`` is the matched material number, or a readable phrase such as
    ``UPC 123456789 not found`` for unmatched items.
    """
    return f"{label}, {format_quantity(qty)}, {format_money(total)}"


def style_result_sheet(ws: Worksheet, n_data_rows: int) -> None:
    """Apply professional-but-simple formatting to a freshly written sheet.

    Assumes row 1 is the header and rows 2..n_data_rows+1 are data.
    """
    n_cols = len(RESULT_COLUMNS)
    last_col = get_column_letter(n_cols)
    last_row = n_data_rows + 1

    # Header styling.
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _TOP_WRAP_HEADER

    # Body: wrap + top-align every cell.
    for r in range(2, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).alignment = _TOP_WRAP

    # Column widths.
    for idx, name in enumerate(RESULT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COLUMN_WIDTHS.get(name, 18)

    # Freeze the header row and add an autofilter over the whole table.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{max(last_row, 1)}"

    # Highlight Manual Review rows (Status is column E) via conditional formatting.
    if n_data_rows > 0:
        status_col = get_column_letter(RESULT_COLUMNS.index("Status") + 1)
        rule = FormulaRule(
            formula=[f'${status_col}2="{STATUS_REVIEW}"'],
            fill=_REVIEW_FILL,
            stopIfTrue=False,
        )
        ws.conditional_formatting.add(f"A2:{last_col}{last_row}", rule)
