"""Workbook export: preserve originals, create / number the results sheet."""
import io

import openpyxl

from src import excel_exporter
from src.formatting import RESULT_COLUMNS
from src.models import STATUS_PROCESSED, STATUS_REVIEW, DebitMemoResult


def _source_workbook(extra_sheets=None):
    """A small short-paid-like workbook with a value, a formula, and formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "06.08.26"
    ws["A1"] = "FULL DEBIT DESCRIPTION"
    ws["B1"] = "Amount"
    ws["A2"] = 90091172
    ws["B2"] = 100
    ws["B3"] = "=B2*2"          # formula must survive
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.column_dimensions["A"].width = 30
    for name in extra_sheets or []:
        wb.create_sheet(name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _results():
    return [
        DebitMemoResult(source_file="a.pdf", vendor_reference="90091172",
                        debit_number="1709032065", po_number="4878948624",
                        cell_text="10057258, 70, $6,375.60", status=STATUS_PROCESSED),
        DebitMemoResult(source_file="b.pdf", vendor_reference="90157331",
                        debit_number="1709032066", po_number="4878948625",
                        cell_text="UPC 123 not found, 5, $1.00", status=STATUS_REVIEW,
                        review_reason="UPC 123 not found in material list"),
    ]


# (15) Original worksheets remain unchanged.
def test_originals_unchanged():
    src = _source_workbook()
    out = excel_exporter.build_workbook(src, _results())

    before = openpyxl.load_workbook(io.BytesIO(src))
    after = openpyxl.load_workbook(io.BytesIO(out))
    ws_b, ws_a = before["06.08.26"], after["06.08.26"]
    assert ws_a["A1"].value == ws_b["A1"].value == "FULL DEBIT DESCRIPTION"
    assert ws_a["A2"].value == 90091172
    assert ws_a["B3"].value == "=B2*2"               # formula preserved
    assert ws_a["A1"].font.bold is True              # formatting preserved
    assert ws_a.column_dimensions["A"].width == 30   # width preserved


# (16) New results sheet creation with the right columns and rows.
def test_results_sheet_created():
    out = excel_exporter.build_workbook(_source_workbook(), _results())
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb.sheetnames[-1] == "Debit Memo Results"
    ws = wb["Debit Memo Results"]
    assert [c.value for c in ws[1]] == RESULT_COLUMNS
    assert ws["A2"].value == "90091172"
    assert ws["D2"].value == "10057258, 70, $6,375.60"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    # Manual-review highlight present as a conditional-formatting rule.
    assert len(ws.conditional_formatting._cf_rules) >= 1


# (17) Existing results-sheet name collision -> next numbered name.
def test_results_sheet_name_collision():
    src = _source_workbook(extra_sheets=["Debit Memo Results", "Debit Memo Results 2"])
    out = excel_exporter.build_workbook(src, _results())
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert "Debit Memo Results 3" in wb.sheetnames
    # Pre-existing sheets are not overwritten.
    assert "Debit Memo Results" in wb.sheetnames
    assert "Debit Memo Results 2" in wb.sheetnames


def test_unique_sheet_name_helper():
    assert excel_exporter.unique_sheet_name([]) == "Debit Memo Results"
    assert excel_exporter.unique_sheet_name(["Debit Memo Results"]) == "Debit Memo Results 2"
    assert excel_exporter.unique_sheet_name(
        ["Debit Memo Results", "Debit Memo Results 2"]) == "Debit Memo Results 3"
